# pip install setuptools (in venv)
# add pandas to project
# add openpyxl to project

import pandas as pd
import openpyxl as xl
from openpyxl.chart import BarChart, Reference
import xlwings as xw


# PANDAS
dataframe1 = pd.read_excel('exceltest.xlsx')

print(dataframe1)


# OPENPYXL
dataframe = xl.load_workbook('exceltest.xlsx')
dataframe2 = dataframe.active

# Iterate the loop to read the cell values
for row in range(0, dataframe2.max_row):
    for col in dataframe2.iter_cols(1, dataframe2.max_column):
        print(col[row].value)



# # XLWINGS
# # Specifying a sheet
# ws = xw.Book('exceltest.xlsx').sheets['Sheet1']
#
# # Selecting data from
# # a single cell
# v1 = ws.range("A1:A7").value
# # v2 = ws.range("F5").value
# print("Result:", v1, v2)




# EXAMPLE
wb = xl.load_workbook('exceltest.xlsx')
sheet = wb['Sheet1']

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    if cell.value is None:
        print('None type')
    else:
        corrected_price = float(cell.value.replace('$', '')) * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

wb.save('python-spreadsheet2.xlsx')


